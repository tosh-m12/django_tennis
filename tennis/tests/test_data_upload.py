"""
データ集計表アップロード（差分プレビュー→確定→反映）のテスト。
ダウンロード→編集→アップロードの往復を実ファイルで検証する。
"""
from __future__ import annotations

import datetime
import io

from django.test import TestCase, override_settings
from django.urls import reverse

from openpyxl import load_workbook, Workbook

from tennis.models import EventParticipant, ParticipantFlag

from .factories import (
    make_club,
    make_event,
    make_member,
    make_ep,
    make_club_flag,
    make_participant_flag,
)

_NO_MANIFEST_STORAGES = override_settings(
    STORAGES={
        "staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"},
    }
)

PERIOD = {"start": "2026-05-01", "end": "2026-05-31"}


@_NO_MANIFEST_STORAGES
class DataUploadRoundTripTests(TestCase):
    def setUp(self):
        self.club = make_club()
        self.ev = make_event(self.club, date=datetime.date(2026, 5, 1), title="練習")
        self.m = make_member(self.club, "山田", member_no=1)
        self.ep = make_ep(self.ev, member=self.m, attendance="yes")
        self.flag = make_club_flag(self.club, "車", 1)  # check モード

    def _download_xlsx(self):
        url = reverse("tennis:club_data_download", args=[self.club.public_token, self.club.admin_token])
        resp = self.client.get(url, {**PERIOD, "format": "xlsx"})
        self.assertEqual(resp.status_code, 200)
        return load_workbook(io.BytesIO(resp.content))

    def _edit_attendance(self, wb, row_key, new_jp):
        ws = wb["出欠"]
        machine = [c.value for c in ws[2]]
        col = machine.index(f"event:{self.ev.id}") + 1
        for r in range(4, ws.max_row + 1):
            if ws.cell(row=r, column=1).value == row_key:
                ws.cell(row=r, column=col).value = new_jp
                return
        raise AssertionError("row not found")

    def _to_bytes(self, wb):
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _upload(self, content, name="edit.xlsx"):
        url = reverse("tennis:club_data_upload", args=[self.club.public_token, self.club.admin_token])
        f = io.BytesIO(content)
        f.name = name
        return self.client.post(url, {"file": f})

    def _apply(self):
        url = reverse("tennis:club_data_apply", args=[self.club.public_token, self.club.admin_token])
        return self.client.post(url, {})

    def test_attendance_round_trip(self):
        wb = self._download_xlsx()
        self._edit_attendance(wb, f"m:{self.m.id}", "不参加")
        resp = self._upload(self._to_bytes(wb))
        self.assertEqual(resp.status_code, 200)
        result = resp.context["result"]
        self.assertEqual(result["errors"], [])
        self.assertEqual(len(result["changes"]), 1)
        self.assertEqual(result["changes"][0]["new_code"], "no")
        # まだDBは変わっていない
        self.ep.refresh_from_db()
        self.assertEqual(self.ep.attendance, "yes")
        # 確定で反映
        self._apply()
        self.ep.refresh_from_db()
        self.assertEqual(self.ep.attendance, "no")

    def test_bad_value_blocks_apply(self):
        wb = self._download_xlsx()
        self._edit_attendance(wb, f"m:{self.m.id}", "出る")  # 不正値
        resp = self._upload(self._to_bytes(wb))
        result = resp.context["result"]
        self.assertTrue(result["errors"])
        self.assertFalse(resp.context["can_apply"])
        # applyしてもセッションに無いので反映されない
        self._apply()
        self.ep.refresh_from_db()
        self.assertEqual(self.ep.attendance, "yes")

    def test_flag_round_trip(self):
        wb = self._download_xlsx()
        # 共通_車 シートで 山田 の該当イベント列に ✓
        ws = wb["共通_車"]
        machine = [c.value for c in ws[2]]
        col = machine.index(f"event:{self.ev.id}") + 1
        for r in range(4, ws.max_row + 1):
            if ws.cell(row=r, column=1).value == f"m:{self.m.id}":
                ws.cell(row=r, column=col).value = "✓"
        resp = self._upload(self._to_bytes(wb))
        result = resp.context["result"]
        self.assertEqual(result["errors"], [])
        self.assertEqual(len(result["changes"]), 1)
        self._apply()
        pf = ParticipantFlag.objects.get(event_participant=self.ep, club_flag_definition=self.flag)
        self.assertTrue(pf.is_on)

    def test_wrong_club_file_rejected(self):
        other = make_club()
        wb = self._download_xlsx()
        # meta の club_id を別クラブに書き換え
        meta = wb["メタ"]
        for r in range(1, meta.max_row + 1):
            if meta.cell(row=r, column=1).value == "club_id":
                meta.cell(row=r, column=2).value = str(other.id)
        resp = self._upload(self._to_bytes(wb))
        self.assertTrue(resp.context["result"]["errors"])

    def test_snapshot_mismatch_blocks_apply(self):
        wb = self._download_xlsx()
        self._edit_attendance(wb, f"m:{self.m.id}", "不参加")
        self._upload(self._to_bytes(wb))  # プレビューOK・session保持
        # 確定前に別の変更でDBが動く → snapshot不一致
        self.ep.attendance = "maybe"
        self.ep.save(update_fields=["attendance", "updated_at"])
        resp = self._apply()
        self.assertIn("再ダウンロード", resp.context["error"])
        self.ep.refresh_from_db()
        self.assertEqual(self.ep.attendance, "maybe")  # 反映されない

    def test_non_xlsx_garbage_is_parse_error(self):
        resp = self._upload(b"not a real file", name="x.xlsx")
        # パースエラー → error 表示（500にしない）
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(resp.context["error"])

"""
データ集計表のダウンロード（縦持ち long 形式・Excel/CSV）のテスト。
"""
from __future__ import annotations

import datetime
import io
import zipfile

from django.test import TestCase, override_settings
from django.urls import reverse

from openpyxl import load_workbook

from .factories import (
    make_club,
    make_event,
    make_member,
    make_ep,
    make_club_flag,
    make_participant_flag,
)

_NO_MANIFEST_STORAGES = override_settings(
    STORAGES={
        "staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"},
    }
)


@_NO_MANIFEST_STORAGES
class DataDownloadTests(TestCase):
    def setUp(self):
        self.club = make_club()
        self.ev = make_event(self.club, date=datetime.date(2026, 5, 1), title="練習会")
        self.m = make_member(self.club, "山田", member_no=1)
        self.ep = make_ep(self.ev, member=self.m, attendance="yes")
        self.flag = make_club_flag(self.club, "車", 1)
        make_participant_flag(self.ep, club_flag=self.flag, is_on=True)
        self.params = {"start": "2026-05-01", "end": "2026-05-31"}

    def _url(self, fmt):
        url = reverse("tennis:club_data_download", args=[self.club.public_token, self.club.admin_token])
        return url, {**self.params, "format": fmt}

    def test_admin_token_required(self):
        url = reverse("tennis:club_data_download", args=[self.club.public_token, "wrong"])
        self.assertEqual(self.client.get(url, self.params).status_code, 400)

    def test_xlsx_is_single_long_sheet(self):
        url, p = self._url("xlsx")
        resp = self.client.get(url, p)
        self.assertEqual(resp.status_code, 200)
        wb = load_workbook(io.BytesIO(resp.content))
        # シートは メタ＋データ の2枚だけ（フラグごとの別シートは無い）
        self.assertEqual(set(wb.sheetnames), {"メタ", "データ"})
        ws = wb["データ"]
        header = [c.value for c in ws[1]]
        self.assertEqual(header, ["row_key", "名前", "event", "イベント", "item", "項目", "値"])
        body = [[c.value for c in row] for row in ws.iter_rows(min_row=2)]
        # 出欠の行
        att = next(r for r in body if r[0] == f"m:{self.m.id}" and r[4] == "attendance")
        self.assertEqual(att[2], f"event:{self.ev.id}")
        self.assertEqual(att[6], "参加")
        # フラグの行（同じ1枚に積まれている）
        flg = next(r for r in body if r[0] == f"m:{self.m.id}" and r[4] == f"clubflag:{self.flag.id}")
        self.assertEqual(flg[6], "✓")

    def test_csv_zip_two_files(self):
        url, p = self._url("csv")
        resp = self.client.get(url, p)
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp["Content-Type"], "application/zip")
        zf = zipfile.ZipFile(io.BytesIO(resp.content))
        names = zf.namelist()
        self.assertIn("00_meta.csv", names)
        self.assertIn("01_データ.csv", names)
        meta = zf.read("00_meta.csv").decode("utf-8-sig")
        self.assertIn("snapshot_token", meta)

    def test_snapshot_token_changes_when_data_changes(self):
        from tennis.views import build_club_data_matrices
        t1 = build_club_data_matrices(
            self.club, datetime.date(2026, 5, 1), datetime.date(2026, 5, 31)
        )["snapshot_token"]
        self.ep.attendance = "no"
        self.ep.save(update_fields=["attendance", "updated_at"])
        t2 = build_club_data_matrices(
            self.club, datetime.date(2026, 5, 1), datetime.date(2026, 5, 31)
        )["snapshot_token"]
        self.assertNotEqual(t1, t2)

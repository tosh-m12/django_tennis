"""
データ集計表アップロードのパース・検証・差分・適用（フェーズ1b：出欠＋フラグ値）。

安全方針:
- まず全件を検証し、ブロッキングエラーが1つでもあれば一切適用しない。
- 適用は transaction.atomic で一括。
- snapshot トークンでダウンロード後の DB 変更を検知（confirm 時にズレていれば中止）。
- 値はホワイトリスト。行キー/列キーは安定 ID で突合（名前依存しない＝メンバー）。
- 人の新規追加はしない（既存メンバーの参加記録作成は可。ゲストは既存記録のみ更新）。
- 名簿シートの編集（固定/順/統合/削除）はこのフェーズでは無視し、注記する。

フォーマット: 各データシートは
  行1 = ["#meta", kind, ref_id]
  行2 = 機械ヘッダ（__rowkey__, __name__, event:<id> / eventflag:<id> ...）
  行3 = 人間ヘッダ
  行4〜 = データ（row_key, 表示名, 値...）
"""
from __future__ import annotations

import csv
import io
import zipfile

from openpyxl import load_workbook

from .models import (
    Event,
    EventParticipant,
    Member,
    ParticipantFlag,
    ClubFlagDefinition,
    EventFlagDefinition,
)
from .data_export import ATT_JP_TO_CODE, FORMAT_VERSION

SUPPORTED_VERSIONS = {"1"}
MAX_CELLS = 200000  # サイズ上限（行×列の総セル数の目安）

ON_TOKENS = {"✓", "✔", "1", "on", "true", "はい", "○", "〇"}
OFF_TOKENS = {"", "0", "false", "いいえ", "×", "x"}


class ParseError(Exception):
    pass


# ============================================================
# パース
# ============================================================

def parse_upload(filename, content):
    """アップロードファイルを共通中間表現にパース。xlsx か csv-zip を判定。"""
    name = (filename or "").lower()
    if name.endswith(".xlsx"):
        return _parse_xlsx(content)
    if name.endswith(".zip"):
        return _parse_csv_zip(content)
    # 拡張子不明：中身で試行
    try:
        return _parse_xlsx(content)
    except Exception:
        return _parse_csv_zip(content)


def _norm(v):
    if v is None:
        return ""
    return str(v).strip()


def _parse_xlsx(content):
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        raise ParseError(f"Excelとして読めません: {e}")
    meta = {}
    sheets = []
    for ws in wb.worksheets:
        rows = [[_norm(c) for c in r] for r in ws.iter_rows(values_only=True)]
        if ws.title == "メタ" or (rows and rows[0] and rows[0][0] == "format_version"):
            meta = _parse_meta_rows(rows)
            continue
        if not rows or not rows[0] or rows[0][0] != "#meta":
            continue  # データシートでない
        sheets.append(_sheet_from_rows(rows))
    return {"meta": meta, "sheets": sheets}


def _parse_csv_zip(content):
    try:
        zf = zipfile.ZipFile(io.BytesIO(content))
    except Exception as e:
        raise ParseError(f"ZIP(CSV束)として読めません: {e}")
    meta = {}
    sheets = []
    for fn in sorted(zf.namelist()):
        if not fn.lower().endswith(".csv"):
            continue
        text = zf.read(fn).decode("utf-8-sig", errors="replace")
        rows = [[_norm(c) for c in r] for r in csv.reader(io.StringIO(text))]
        if not rows:
            continue
        if rows[0] and rows[0][0] == "format_version":
            meta = _parse_meta_rows(rows)
            continue
        if rows[0] and rows[0][0] == "#meta":
            sheets.append(_sheet_from_rows(rows))
    return {"meta": meta, "sheets": sheets}


def _parse_meta_rows(rows):
    meta = {}
    for r in rows:
        if not r or not r[0] or r[0].startswith("#"):
            continue
        key = r[0]
        val = r[1] if len(r) > 1 else ""
        if key in ("format_version", "club_id", "club_name", "period_start",
                   "period_end", "generated_at", "snapshot_token"):
            meta[key] = val
    return meta


def _sheet_from_rows(rows):
    kind = rows[0][1] if len(rows[0]) > 1 else ""
    ref_id = rows[0][2] if len(rows[0]) > 2 else ""
    machine = rows[1] if len(rows) > 1 else []
    data_rows = rows[3:] if len(rows) > 3 else []
    return {"kind": kind, "ref_id": ref_id, "machine_header": machine, "data_rows": data_rows}


# ============================================================
# 検証＋差分
# ============================================================

def analyze(club, parsed):
    """
    パース結果を検証し、差分（適用予定の変更）を返す。
    戻り値 dict:
      errors[], warnings[], changes[], ignored[],
      file_snapshot, current_snapshot, snapshot_match,
      period_start, period_end, summary{}
    """
    from .views import build_club_data_matrices  # 遅延 import（循環回避）

    errors, warnings, ignored, changes = [], [], [], []
    meta = parsed.get("meta", {})

    if meta.get("format_version") not in SUPPORTED_VERSIONS:
        errors.append(f"未対応のフォーマット版です（{meta.get('format_version')!r}）。")
    if str(meta.get("club_id")) != str(club.id):
        errors.append("別のクラブのファイルです（club_id 不一致）。")
        return _result(errors, warnings, changes, ignored, meta, None)

    # 対象イベント・メンバー・フラグ定義を読み込み
    events = {e.id: e for e in Event.objects.filter(club=club)}
    members = {m.id: m for m in Member.objects.filter(club=club)}
    club_flags = {f.id: f for f in ClubFlagDefinition.objects.filter(club=club)}
    event_flags = {f.id: f for f in EventFlagDefinition.objects.filter(event__club=club)}

    # 現在値の参照（EP と ParticipantFlag）
    eps = list(
        EventParticipant.objects.filter(event__club=club).select_related("member")
    )
    ep_by_member_event = {}      # (member_id, event_id) -> ep
    ep_by_guest_event = {}       # (name, event_id) -> ep
    for ep in eps:
        if ep.member_id:
            ep_by_member_event[(ep.member_id, ep.event_id)] = ep
        else:
            nm = (ep.display_name or "").strip()
            ep_by_guest_event.setdefault((nm, ep.event_id), ep)
    pf_by_ep = {}                # (ep_id, scope, flag_id) -> ParticipantFlag
    for pf in ParticipantFlag.objects.filter(event_participant__event__club=club):
        if pf.club_flag_definition_id:
            pf_by_ep[(pf.event_participant_id, "club", pf.club_flag_definition_id)] = pf
        elif pf.event_flag_definition_id:
            pf_by_ep[(pf.event_participant_id, "event", pf.event_flag_definition_id)] = pf

    def resolve_row(row_key):
        """row_key -> (member or None, display_name, error_or_None)"""
        if row_key.startswith("m:"):
            try:
                mid = int(row_key[2:])
            except ValueError:
                return None, "", f"行キー不正: {row_key}"
            m = members.get(mid)
            if not m:
                return None, "", f"存在しないメンバー: {row_key}"
            return m, m.display_name, None
        if row_key.startswith("g:"):
            return None, row_key[2:], None
        return None, "", f"行キー不正: {row_key}"

    def current_ep(member, name, event_id):
        if member:
            return ep_by_member_event.get((member.id, event_id))
        return ep_by_guest_event.get((name, event_id))

    total_cells = 0
    for sheet in parsed.get("sheets", []):
        kind = sheet["kind"]
        machine = sheet["machine_header"]
        if kind == "roster":
            # 名簿の編集（固定/順/統合/削除）は今フェーズ未対応
            for dr in sheet["data_rows"]:
                pass
            ignored.append("名簿シートの編集（固定/表示順/統合/削除）は現バージョンでは反映されません。")
            continue
        if kind not in ("attendance", "clubflag", "eventflags"):
            continue

        # 列の解決
        col_targets = []  # (col_index, ("event", event_id) | ("flag", scope, flag_def))
        for ci, key in enumerate(machine):
            if ci < 2:
                continue
            if key.startswith("event:"):
                try:
                    ev_id = int(key.split(":", 1)[1])
                except ValueError:
                    errors.append(f"列キー不正: {key}")
                    continue
                if ev_id not in events:
                    errors.append(f"存在しないイベント列: {key}")
                    continue
                col_targets.append((ci, ("event", ev_id)))
            elif key.startswith("eventflag:"):
                try:
                    fid = int(key.split(":", 1)[1])
                except ValueError:
                    errors.append(f"列キー不正: {key}")
                    continue
                fdef = event_flags.get(fid)
                if not fdef:
                    errors.append(f"存在しない固有フラグ列: {key}")
                    continue
                col_targets.append((ci, ("flag", "event", fdef)))

        # clubflag シートはフラグ定義が ref_id、列はイベント
        club_fdef = None
        if kind == "clubflag":
            try:
                club_fdef = club_flags.get(int(sheet["ref_id"]))
            except (TypeError, ValueError):
                club_fdef = None
            if not club_fdef:
                errors.append(f"存在しない共通フラグ: ref={sheet['ref_id']}")
                continue

        for dr in sheet["data_rows"]:
            if not dr or not dr[0]:
                continue
            row_key = dr[0]
            member, name, rerr = resolve_row(row_key)
            if rerr:
                errors.append(rerr)
                continue
            for ci, target in col_targets:
                total_cells += 1
                raw = dr[ci] if ci < len(dr) else ""
                if kind == "attendance":
                    _diff_attendance(raw, member, name, target[1], events,
                                     current_ep, changes, errors)
                elif kind == "clubflag":
                    _diff_flag(raw, member, name, target[1], "club", club_fdef, events,
                               current_ep, pf_by_ep, changes, errors)
                else:  # eventflags
                    _, scope, fdef = target
                    _diff_flag(raw, member, name, fdef.event_id, "event", fdef, events,
                               current_ep, pf_by_ep, changes, errors)

    if total_cells > MAX_CELLS:
        errors.append("ファイルが大きすぎます。")

    # snapshot（楽観ロック）
    file_snapshot = meta.get("snapshot_token", "")
    current_snapshot = None
    ps, pe = meta.get("period_start"), meta.get("period_end")
    if ps and pe:
        from datetime import date
        try:
            sd = date.fromisoformat(ps)
            ed = date.fromisoformat(pe)
            current_snapshot = build_club_data_matrices(club, sd, ed)["snapshot_token"]
        except Exception:
            current_snapshot = None
    snapshot_match = bool(file_snapshot and current_snapshot and file_snapshot == current_snapshot)
    if current_snapshot and not snapshot_match:
        warnings.append("ダウンロード後にデータが変更されています。最新を再ダウンロードして編集し直すことを推奨します。")

    return _result(errors, warnings, changes, ignored, meta, current_snapshot, snapshot_match)


def _result(errors, warnings, changes, ignored, meta, current_snapshot, snapshot_match=False):
    return {
        "errors": errors,
        "warnings": warnings,
        "changes": changes,
        "ignored": sorted(set(ignored)),
        "file_snapshot": meta.get("snapshot_token", ""),
        "current_snapshot": current_snapshot,
        "snapshot_match": snapshot_match,
        "period_start": meta.get("period_start"),
        "period_end": meta.get("period_end"),
        "summary": {
            "changes": len(changes),
            "errors": len(errors),
        },
    }


def _diff_attendance(raw, member, name, event_id, events, current_ep, changes, errors):
    if raw not in ATT_JP_TO_CODE:
        errors.append(f"出欠の値が不正です: {raw!r}（参加/不参加/未定/空 のみ）")
        return
    new_code = ATT_JP_TO_CODE[raw]
    ep = current_ep(member, name, event_id)
    cur_code = (ep.attendance if ep else None) or None
    if cur_code == new_code:
        return
    if not member and not ep:
        return  # ゲストで記録が無いセルは作成しない（無視）
    changes.append({
        "kind": "attendance",
        "member_id": member.id if member else None,
        "display_name": name,
        "event_id": event_id,
        "old_text": ATT_CODE_LABEL(cur_code),
        "new_text": ATT_CODE_LABEL(new_code),
        "new_code": new_code,
    })


def _diff_flag(raw, member, name, event_id, scope, fdef, events, current_ep, pf_by_ep, changes, errors):
    # 値の正規化＋検証
    if fdef.input_mode == "digit":
        if raw == "":
            new_is_on, new_val = False, None
        elif raw.isdigit() and len(raw) == 1:
            new_is_on, new_val = True, int(raw)
        else:
            errors.append(f"数値フラグ「{fdef.name}」の値が不正: {raw!r}（0-9 か空）")
            return
    else:  # check
        low = raw.lower()
        if low in ON_TOKENS or raw in ON_TOKENS:
            new_is_on, new_val = True, None
        elif low in OFF_TOKENS or raw in OFF_TOKENS:
            new_is_on, new_val = False, None
        else:
            errors.append(f"チェックフラグ「{fdef.name}」の値が不正: {raw!r}（✓ か空）")
            return

    ep = current_ep(member, name, event_id)
    pf = pf_by_ep.get((ep.id, scope, fdef.id)) if ep else None
    cur_is_on = bool(pf.is_on) if pf else False
    cur_val = pf.value if pf else None
    if cur_is_on == new_is_on and cur_val == new_val:
        return
    if not member and not ep:
        return  # ゲストで記録が無いセルは無視

    changes.append({
        "kind": "flag",
        "scope": scope,
        "flag_id": fdef.id,
        "flag_name": fdef.name,
        "input_mode": fdef.input_mode,
        "member_id": member.id if member else None,
        "display_name": name,
        "event_id": event_id,
        "old_text": _flag_label(fdef.input_mode, cur_is_on, cur_val),
        "new_text": _flag_label(fdef.input_mode, new_is_on, new_val),
        "new_is_on": new_is_on,
        "new_val": new_val,
    })


def ATT_CODE_LABEL(code):
    return {"yes": "参加", "no": "不参加", "maybe": "未定", None: "（空）"}.get(code, "（空）")


def _flag_label(input_mode, is_on, val):
    if input_mode == "digit":
        return "（空）" if val is None else str(val)
    return "✓" if is_on else "（空）"


# ============================================================
# 適用
# ============================================================

def apply_changes(club, changes, actor_kind="admin"):
    """検証済み changes を atomic に適用。AuditLog を残す。適用件数を返す。"""
    from django.db import transaction
    from .models import AuditLog
    from .views import _get_or_create_ep

    applied = 0
    with transaction.atomic():
        for ch in changes:
            event = Event.objects.select_for_update().get(id=ch["event_id"], club=club)
            ep = _resolve_ep_for_apply(event, ch, _get_or_create_ep)
            if ep is None:
                continue
            if ch["kind"] == "attendance":
                _apply_attendance(ep, ch["new_code"])
            else:
                _apply_flag(ep, ch)
            applied += 1

        AuditLog.objects.create(
            club=club,
            actor_token_kind=actor_kind,
            action="data_upload_apply",
            payload_json={
                "applied": applied,
                "by_kind": _count_by_kind(changes),
            },
        )
    return applied


def _resolve_ep_for_apply(event, ch, get_or_create_ep):
    if ch.get("member_id"):
        member = Member.objects.get(id=ch["member_id"], club=event.club)
        return get_or_create_ep(event, member, member.display_name)
    # ゲスト：既存記録のみ（新規作成しない）
    return (
        EventParticipant.objects
        .filter(event=event, member__isnull=True, display_name=ch["display_name"])
        .order_by("id")
        .first()
    )


def _apply_attendance(ep, new_code):
    old = ep.attendance or ""
    ep.attendance = new_code or None
    if (new_code or "") != "yes":
        ep.participates_match = False
    elif old != "yes":
        ep.participates_match = True
    ep.save(update_fields=["attendance", "participates_match", "updated_at"])


def _apply_flag(ep, ch):
    scope = ch["scope"]
    defaults = {"is_on": ch["new_is_on"], "value": ch["new_val"]}
    if scope == "club":
        obj, _ = ParticipantFlag.objects.get_or_create(
            event_participant=ep, club_flag_definition_id=ch["flag_id"], defaults=defaults
        )
    else:
        obj, _ = ParticipantFlag.objects.get_or_create(
            event_participant=ep, event_flag_definition_id=ch["flag_id"], defaults=defaults
        )
    obj.is_on = ch["new_is_on"]
    obj.value = ch["new_val"]
    obj.save(update_fields=["is_on", "value", "updated_at"])


def _count_by_kind(changes):
    out = {}
    for ch in changes:
        k = ch["kind"]
        out[k] = out.get(k, 0) + 1
    return out

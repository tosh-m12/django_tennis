"""
データ集計表＋名簿の Excel(.xlsx) / CSV束(ZIP) 出力。

フォーマット（v1）— アップロード時に機械可読で同定できるよう、各データシートは
2行ヘッダ（1行目=機械キー、2行目=人間ラベル）、データは3行目以降。
シートの種別と参照先IDは「メタ」シート（CSVは meta.csv）のレジストリに記録する。

行キー: メンバー=`m:<member_id>` / ゲスト=`g:<display_name>`
列キー（出欠/共通フラグ）: `event:<event_id>`
列キー（固有フラグ）: `eventflag:<flag_id>`
"""
from __future__ import annotations

import csv
import io
import zipfile

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

FORMAT_VERSION = "1"

# 出欠コード ↔ 日本語ラベル（人間が編集しやすいよう日本語で出す）
ATT_CODE_TO_JP = {"yes": "参加", "no": "不参加", "maybe": "未定", None: ""}
ATT_JP_TO_CODE = {"参加": "yes", "不参加": "no", "未定": "maybe", "": None}


def _row_key(row):
    """マトリクスの row（build_club_data_matrices）から安定キー文字列を作る。"""
    if row.get("member_id"):
        return f"m:{row['member_id']}"
    return f"g:{row['display_name']}"


def _event_label(ev):
    label = ev.date.strftime("%-m/%-d")
    if ev.title:
        label += f" {ev.title}"
    return label


def _safe_sheet_name(base, used):
    """Excel シート名の制約（31字・記号禁止・一意）に収める。"""
    for ch in "[]:*?/\\":
        base = base.replace(ch, "_")
    base = base.strip() or "sheet"
    name = base[:31]
    i = 2
    while name in used:
        suffix = f"_{i}"
        name = base[:31 - len(suffix)] + suffix
        i += 1
    used.add(name)
    return name


def build_sheets(club, start_d, end_d, data):
    """
    出力に必要なシート定義のリストを返す（Excel/CSV 共通の中間表現）。
    各要素: {"name", "kind", "ref_id", "machine_header", "human_header", "data_rows"}
    先頭は必ずメタ情報（kind="meta"）。
    """
    rows = data["rows"]
    events = data["events"]

    sheets = []

    # --- 出欠シート ---
    machine = ["__rowkey__", "__name__"] + [f"event:{ev.id}" for ev in events]
    human = ["キー", "メンバー"] + [_event_label(ev) for ev in events]
    data_rows = []
    for r in data["attendance_table"]:
        rk = _row_key(r["row"])
        cells = [ATT_CODE_TO_JP.get(c["attendance"], "") for c in r["cells"]]
        data_rows.append([rk, r["row"]["display_name"]] + cells)
    sheets.append({
        "name": "出欠", "kind": "attendance", "ref_id": "",
        "machine_header": machine, "human_header": human, "data_rows": data_rows,
    })

    # --- 共通フラグシート（フラグごと）---
    for t in data["club_flag_tables"]:
        f = t["flag"]
        machine = ["__rowkey__", "__name__"] + [f"event:{ev.id}" for ev in events]
        human = ["キー", "メンバー"] + [_event_label(ev) for ev in events]
        data_rows = []
        for r in t["rows"]:
            rk = _row_key(r["row"])
            cells = [c["text"] for c in r["cells"]]
            data_rows.append([rk, r["row"]["display_name"]] + cells)
        sheets.append({
            "name": f"共通_{f.name}", "kind": "clubflag", "ref_id": str(f.id),
            "machine_header": machine, "human_header": human, "data_rows": data_rows,
        })

    # --- 固有フラグシート（イベントごと・列=フラグ）---
    for blk in data["event_flag_blocks"]:
        ev = blk["event"]
        flags = blk["flags"]
        machine = ["__rowkey__", "__name__"] + [f"eventflag:{f.id}" for f in flags]
        human = ["キー", "メンバー"] + [f.name for f in flags]
        data_rows = []
        for r in blk["rows"]:
            rk = _row_key(r["row"])
            cells = [c["text"] for c in r["cells"]]
            data_rows.append([rk, r["row"]["display_name"]] + cells)
        sheets.append({
            "name": f"固有_{_event_label(ev)}", "kind": "eventflags", "ref_id": str(ev.id),
            "machine_header": machine, "human_header": human, "data_rows": data_rows,
        })

    # --- 名簿シート（参照用。編集は将来フェーズ）---
    machine = ["__rowkey__", "display_name", "is_fixed", "member_no", "merge_into", "op"]
    human = ["キー", "表示名", "固定", "表示順", "統合先", "操作"]
    data_rows = []
    for r in rows:
        rk = _row_key(r)
        fixed = "固定" if r.get("is_fixed") else ""
        mno = r.get("member_no") if r.get("member_no") is not None else ""
        data_rows.append([rk, r["display_name"], fixed, mno, "", ""])
    sheets.append({
        "name": "名簿", "kind": "roster", "ref_id": "",
        "machine_header": machine, "human_header": human, "data_rows": data_rows,
    })

    return sheets


def _meta_rows(club, start_d, end_d, generated_at, snapshot_token, sheets):
    """メタ情報（key/value ＋ シートレジストリ）の行リスト。"""
    meta = [
        ["format_version", FORMAT_VERSION],
        ["club_id", str(club.id)],
        ["club_name", club.name],
        ["period_start", start_d.isoformat()],
        ["period_end", end_d.isoformat()],
        ["generated_at", generated_at],
        ["snapshot_token", snapshot_token],
        ["", ""],
        ["# sheet", "kind", "ref_id"],
    ]
    for s in sheets:
        meta.append([s["name"], s["kind"], s["ref_id"]])
    return meta


def build_workbook(club, start_d, end_d, data, generated_at):
    """openpyxl Workbook を組み立てて返す。"""
    sheets = build_sheets(club, start_d, end_d, data)
    wb = Workbook()
    used = set()

    # メタシート（先頭）
    meta_ws = wb.active
    meta_ws.title = _safe_sheet_name("メタ", used)
    for row in _meta_rows(club, start_d, end_d, generated_at, data["snapshot_token"], sheets):
        meta_ws.append(row)

    # データシート
    for s in sheets:
        ws = wb.create_sheet(title=_safe_sheet_name(s["name"], used))
        ws.append(s["machine_header"])
        ws.append(s["human_header"])
        for dr in s["data_rows"]:
            ws.append(dr)
        # 機械ヘッダ行は隠す（人間には2行目を見せる）
        ws.row_dimensions[1].hidden = True
        ws.freeze_panes = "C3"
        # 列幅をざっくり
        ws.column_dimensions[get_column_letter(2)].width = 16

    return wb


def workbook_bytes(club, start_d, end_d, data, generated_at):
    wb = build_workbook(club, start_d, end_d, data, generated_at)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def csv_zip_bytes(club, start_d, end_d, data, generated_at):
    """シートごとに CSV を作り ZIP にまとめて返す。"""
    sheets = build_sheets(club, start_d, end_d, data)
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        # meta.csv
        meta_io = io.StringIO()
        w = csv.writer(meta_io)
        for row in _meta_rows(club, start_d, end_d, generated_at, data["snapshot_token"], sheets):
            w.writerow(row)
        zf.writestr("00_meta.csv", meta_io.getvalue())

        used_files = set()
        for i, s in enumerate(sheets, 1):
            fname_base = f"{i:02d}_{s['name']}"
            for ch in "/\\":
                fname_base = fname_base.replace(ch, "_")
            fname = f"{fname_base}.csv"
            j = 2
            while fname in used_files:
                fname = f"{fname_base}_{j}.csv"
                j += 1
            used_files.add(fname)

            sio = io.StringIO()
            w = csv.writer(sio)
            w.writerow(s["machine_header"])
            w.writerow(s["human_header"])
            for dr in s["data_rows"]:
                w.writerow(dr)
            # CSV は Excel で文字化けしないよう BOM 付き UTF-8
            zf.writestr(fname, "﻿" + sio.getvalue())
    return buf.getvalue()
